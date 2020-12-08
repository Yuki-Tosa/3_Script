name = input('ティッカーシンボルを入力: ')


import openpyxl
import pandas as pd


# バランスシート

in_bsf_workbook = openpyxl.load_workbook('../2_Format/BalanceSheet_Format.xlsx', read_only=True, data_only=True)
in_bsf_sheet = in_bsf_workbook[in_bsf_workbook.sheetnames[0]]

# Sheet.valuesをpd.DataFrameへと変換
bsf_data = in_bsf_sheet.values
bsf_data_list = list(bsf_data)
bsf_df = pd.DataFrame(bsf_data_list[1:], columns=bsf_data_list[0])

# 6行目以降を削除
bsf_df_tmp_1 = bsf_df.iloc[range(5), :]

# 同一名のカラムのうち、一番目を削除し、二番目を残す
bsf_df_tmp_2 = bsf_df_tmp_1.loc[:,~bsf_df_tmp_1.columns.duplicated(keep = "last")]

# 一列目を削除
bsf_df_result = bsf_df_tmp_2.drop(bsf_df_tmp_2.columns[0], axis=1)


# インカムステートメント

in_isf_workbook = openpyxl.load_workbook('../2_Format/IncomeStatement_Format.xlsx', read_only=True, data_only=True)
in_isf_sheet = in_isf_workbook[in_isf_workbook.sheetnames[0]]

# Sheet.valuesをpd.DataFrameへと変換
isf_data = in_isf_sheet.values
isf_data_list = list(isf_data)
isf_df = pd.DataFrame(isf_data_list[1:], columns=isf_data_list[0])

# 7行目以降を削除
isf_df_tmp_1 = isf_df.iloc[range(6), :]

# 同一名のカラムのうち、一番目を削除し、二番目を残す
isf_df_tmp_2 = isf_df_tmp_1.loc[:,~isf_df_tmp_1.columns.duplicated(keep = "last")]

# 一行目と一列目を削除
isf_df_result = isf_df_tmp_2.drop(index=isf_df_tmp_2.index[0], columns=isf_df_tmp_2.columns[0])


# 両者をマージ

name_result = pd.merge(bsf_df_result, isf_df_result, on='Breakdown')

name_result_pass = '../1_Storage/{0}_new_join.xlsx'.format(name)
name_result.to_excel(name_result_pass, index=False)